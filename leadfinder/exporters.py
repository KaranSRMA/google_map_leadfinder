from __future__ import annotations

import json
from dataclasses import asdict, fields
from pathlib import Path

from .models import DependencyError, Lead


def load_openpyxl() -> dict[str, object]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise DependencyError(
            "openpyxl is not installed yet.\n"
            "Install it with:\n"
            "  .\\lead_finder\\Scripts\\python.exe -m pip install openpyxl"
        ) from exc

    return {
        "Workbook": Workbook,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
    }


def export_json(path: Path, leads: list[Lead]) -> None:
    payload = [asdict(lead) for lead in leads]
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)


def export_xlsx(path: Path, leads: list[Lead]) -> None:
    modules = load_openpyxl()
    Workbook = modules["Workbook"]
    Alignment = modules["Alignment"]
    Border = modules["Border"]
    Font = modules["Font"]
    PatternFill = modules["PatternFill"]
    Side = modules["Side"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    headers = [field.name for field in fields(Lead)]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    priority_fills = {
        "High": PatternFill(fill_type="solid", fgColor="FFF2CC"),
        "Medium": PatternFill(fill_type="solid", fgColor="FCE4D6"),
        "Low": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    }
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border

    for lead in leads:
        row = asdict(lead)
        sheet.append([row[header] for header in headers])

    for row_index, lead in enumerate(leads, start=2):
        row_fill = priority_fills.get(lead.priority)
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if row_fill is not None:
                cell.fill = row_fill

    for column_cells in sheet.columns:
        lengths = []
        for cell in column_cells:
            for line in str(cell.value or "").splitlines():
                lengths.append(len(line))
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(lengths + [10]) + 2, 45)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
